import csv
import json
import os
import sys
import time
from pathlib import Path
from datetime import datetime, date

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.support.ui import Select as SeleniumSelect

from gdrive_utils import upload_file

# ==========================================
# ENCODING (deve vir antes de qualquer print)
# ==========================================
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    pass

# CONFIGURAÇÕES GERAIS
# =========================================
PBI_LOGIN_EMAIL = os.environ.get("PBI_LOGIN_EMAIL")
PBI_SENHA = os.environ.get("PBI_SENHA")

URL_INADIMPLENCIA = (
    "https://app.powerbi.com/groups/me/reports/bf194288-6c64-4069-978b-cd7d9ae98f6e/"
    "ReportSection69691588076f691dc558?ctid=d7be86d0-66c8-4589-8c27-9cf5f31d3a1d&experience=power-bi&clientSideAuth=0"
)
URL_RELACIONAMENTO = (
    "https://app.powerbi.com/groups/me/reports/6fbbfd34-7ecf-4a40-bdc4-55910d8bf16e/"
    "ReportSectiona45eb931ee53257081e6?ctid=d7be86d0-66c8-4589-8c27-9cf5f31d3a1d&experience=power-bi&clientSideAuth=0"
)
URL_VENDAS = (
    "https://app.powerbi.com/groups/me/reports/6fbbfd34-7ecf-4a40-bdc4-55910d8bf16e/"
    "ReportSectioneef70c731ada62eed0aa?ctid=d7be86d0-66c8-4589-8c27-9cf5f31d3a1d&experience=power-bi&clientSideAuth=0"
)

FRANQUIAS = ["SALVADOR PARIPE", "SALVADOR LARGO", "SALVADOR CENTRO", "SALVADOR CABULA", "ILHEUS"]
FRANQUIAS_VENDAS = FRANQUIAS + ["ILHEUS"]   # Relatório 3 inclui também Ilhéus
DATA_FILIACAO_INICIO = "04-05-2026"

DOWNLOAD_DIR = str(Path(__file__).parent / "downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
POST_DOWNLOAD_BUFFER = 2

GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_PBI_ID")

# NOTA: a versão original convertia xlsx -> csv abrindo o Excel de verdade
# (win32com), que salvava em ";" (ponto e vírgula) por causa do separador de
# lista regional do Windows em pt-BR, e em cp1252 (ANSI/Windows-1252), que é
# o encoding padrão do "Salvar como CSV" do Excel em pt-BR — NÃO utf-8-sig.
# O backend do RedeService (sp_importacao_formata_campos_CARTAO_TODOS) espera
# esse encoding de largura fixa por byte; utf-8-sig já causou rejeição da
# importação de Inadimplência com "Invalid length parameter passed to the
# LEFT or SUBSTRING function" (multi-byte de acentos + BOM descasam o
# LEFT/SUBSTRING por posição fixa no SQL).
CSV_DELIMITER = ";"
CSV_ENCODING = "cp1252"

RS_URL_IMPORT = "https://cobranca01.redeservice.com.br/cobranca.be.cartaotodos/Home/Login"
RS_LOGIN = os.environ.get("RS_LOGIN")
RS_SENHA = os.environ.get("RS_SENHA")

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}


# ==========================================
# UTILITÁRIOS BASE (WEB E ARQUIVOS)
# =========================================
class WebUtils:
    @staticmethod
    def safe_type(el, text, delay=0.02):
        try: el.click()
        except: pass
        el.send_keys(Keys.CONTROL, "a"); el.send_keys(Keys.BACKSPACE)
        for ch in text:
            el.send_keys(ch); time.sleep(delay)

    @staticmethod
    def safe_type_fast(el, text):
        el.click()
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.BACKSPACE)
        el.send_keys(text)

    @staticmethod
    def js_click(driver, el):
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        try: el.click()
        except: driver.execute_script("arguments[0].click();", el)


class FileProcessor:
    @staticmethod
    def _file_size(path: Path) -> int:
        try: return path.stat().st_size
        except: return -1

    @staticmethod
    def _size_stable(path: Path, checks=2, interval=0.5) -> bool:
        last = FileProcessor._file_size(path)
        if last < 0: return False
        for _ in range(checks):
            time.sleep(interval)
            cur = FileProcessor._file_size(path)
            if cur != last or cur <= 0: return False
            last = cur
        return True

    @staticmethod
    def wait_for_download(download_dir, since_ts=None, timeout=300, ok_exts=(".xlsx",)):
        end = time.time() + timeout
        since_ts = since_ts or time.time()
        download_dir = Path(download_dir)
        while time.time() < end:
            has_temp = any(download_dir.glob("*.crdownload"))
            finals = [p for p in download_dir.glob("*") if p.suffix.lower() in ok_exts and p.stat().st_mtime >= since_ts]
            if finals and not has_temp:
                target = sorted(finals, key=lambda p: p.stat().st_mtime, reverse=True)[0]
                if FileProcessor._size_stable(target, checks=2, interval=0.5):
                    time.sleep(POST_DOWNLOAD_BUFFER)
                    return target
            time.sleep(0.5)
        raise TimeoutException("Download não finalizou dentro do tempo limite.")

    @staticmethod
    def _valor_csv(valor):
        """
        Formata um valor de célula do jeito que o Excel em pt-BR escreveria
        num "Salvar como CSV": data em DD/MM/AAAA e número decimal com
        vírgula.

        ATENÇÃO — NÃO MUDAR O FORMATO DE DATA SEM CONFIRMAR NO GESTÃO.
        Já tentei "melhorar" isso renderizando a data conforme o
        number_format da célula (a coluna "Data filiação" vem com formato
        mm-dd-yy, o que daria "08-17-26"). O RedeService NÃO leu a base
        assim — o formato certo é este aqui, DD/MM/AAAA com barra.
        """
        if valor is None:
            return ""
        if isinstance(valor, datetime):
            if (valor.hour, valor.minute, valor.second) == (0, 0, 0):
                return valor.strftime("%d/%m/%Y")
            return valor.strftime("%d/%m/%Y %H:%M:%S")
        if isinstance(valor, date):
            return valor.strftime("%d/%m/%Y")
        if isinstance(valor, float):
            texto = f"{valor:.10f}".rstrip("0").rstrip(".")
            return texto.replace(".", ",")
        return valor

    @staticmethod
    def contar_linhas(caminho_xlsx: Path, linha_inicio: int = 2) -> int:
        """
        Conta quantas linhas de dado tem uma planilha (ignorando o cabeçalho).

        Só pra alimentar o resumo diário — se der qualquer problema aqui,
        devolve 0 em vez de derrubar a automação inteira: contar linha é
        informativo, não pode ser motivo de falhar uma extração que deu certo.
        (Já derrubou uma execução: em read_only o openpyxl às vezes devolve
        max_row = None e a comparação com int estourava TypeError.)
        """
        wb = None
        try:
            wb = openpyxl.load_workbook(caminho_xlsx, read_only=True)
            ws = wb.active
            total = 0
            for indice, linha in enumerate(ws.iter_rows(values_only=True), start=1):
                if indice < linha_inicio:
                    continue
                if any(v is not None and str(v).strip() != "" for v in linha):
                    total += 1
            return total
        except Exception as e:
            print(f"⚠️ Não consegui contar as linhas de {caminho_xlsx.name}: {e}")
            return 0
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    @staticmethod
    def processar_planilha_inadimplencia(caminho_xlsx: Path) -> tuple[Path, int]:
        """
        Preenche a coluna X com 'Base_<Mês>' e converte para CSV.

        Conversão feita direto via openpyxl (lendo os valores das células e
        escrevendo com csv.writer), sem depender do Microsoft Excel estar
        instalado — necessário porque os runners do GitHub Actions não têm
        Office instalado.
        """
        mes_atual = MESES_PT[datetime.now().month]
        valor_base = f"Base_{mes_atual}"
        COLUNA_X = 24
        LINHA_INICIO = 2

        print(f"\n📝 Editando planilha: {caminho_xlsx.name}")
        wb = openpyxl.load_workbook(caminho_xlsx)
        ws = wb.active

        ultima_linha = ws.max_row
        while ultima_linha > LINHA_INICIO:
            if any(ws.cell(row=ultima_linha, column=c).value is not None for c in range(1, COLUNA_X)):
                break
            ultima_linha -= 1

        LINHA_FIM = ultima_linha
        print(f"✏️ Preenchendo coluna X (linhas {LINHA_INICIO}–{LINHA_FIM}) com '{valor_base}'...")

        for linha in range(LINHA_INICIO, LINHA_FIM + 1):
            ws.cell(row=linha, column=COLUNA_X, value=valor_base)

        wb.save(caminho_xlsx)

        caminho_csv = caminho_xlsx.with_suffix(".csv")
        print(f"📄 Convertendo para CSV (delimitador '{CSV_DELIMITER}')...")

        with open(caminho_csv, "w", newline="", encoding=CSV_ENCODING) as f:
            writer = csv.writer(f, delimiter=CSV_DELIMITER)
            for row in ws.iter_rows(min_row=1, max_row=LINHA_FIM):
                writer.writerow([FileProcessor._valor_csv(cell.value) for cell in row])

        wb.close()
        print(f"✅ CSV gerado: {caminho_csv.name}")

        contagem = max(0, LINHA_FIM - LINHA_INICIO + 1)
        return caminho_csv, contagem


class BrowserFactory:
    @staticmethod
    def create_chrome(download_dir=DOWNLOAD_DIR):
        opts = webdriver.ChromeOptions()
        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            # Sem isso, o runner (locale en-US) faz o Power BI renderizar toda
            # a interface em inglês, e os seletores por texto/data-testid do
            # script — escritos em português — deixam de bater.
            "intl.accept_languages": "pt-BR,pt",
        }
        opts.add_experimental_option("prefs", prefs)
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument("--lang=pt-BR")
        opts.add_argument("--accept-lang=pt-BR,pt")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--log-level=3")
        opts.add_argument("--silent")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--window-size=1920,1080")
        if os.environ.get("PBI_HEADLESS", "true").lower() != "false":
            opts.add_argument("--headless=new")
        opts.page_load_strategy = "eager"

        driver = webdriver.Chrome(options=opts)
        try:
            driver.maximize_window()
        except Exception:
            pass
        return driver


# ==========================================
# AUTOMAÇÃO: POWER BI
# =========================================
class PowerBIBot:
    POPUP_XPATH = "//div[contains(@class,'slicer-dropdown-popup') and contains(@style,'display: block') and .//div[@class='slicerBody' and contains(@aria-label,'Franquia')]]"
    XPATH_DATA_INPUT = "//*[@id='pvExplorationHost']/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[6]/transform/div/div[3]/div/div/visual-modern/div/div/div[2]/div/div[1]/div/div[1]/div//input[contains(@class,'date-slicer-input')]"

    # XPaths para o cabeçalho Email (usados para re-fetch anti-stale)
    _EMAIL_HEADER_XPATHS = [
        "//div[@role='columnheader' and @data-query-ref='FT_FILIADOS_INATIVOS.Email']",
        "//div[@role='columnheader' and contains(normalize-space(.), 'Email')]",
    ]

    def __init__(self, driver):
        self.driver = driver
        # 60s (era 35s): em runners de nuvem sem GPU, o Power BI demora mais
        # para renderizar visuais pesados do que numa máquina local.
        self.wait = WebDriverWait(driver, 60)

    # ── Helper anti-stale: sempre busca o email_header fresco do DOM ───────
    def _fresh_email_header(self):
        for xpath in self._EMAIL_HEADER_XPATHS:
            try:
                return self.driver.find_element(By.XPATH, xpath)
            except Exception:
                continue
        return None

    def login(self):
        try:
            email_pbi = WebDriverWait(self.driver, 6).until(EC.element_to_be_clickable((By.ID, "email")))
            WebUtils.safe_type_fast(email_pbi, PBI_LOGIN_EMAIL)
            # O texto do botão varia com o idioma do navegador ("Enviar" em
            # pt-BR, "Submit" em en-US) — tenta os dois, e cai para ENTER no
            # campo se nenhum botão for encontrado a tempo (não depende de
            # idioma nenhum).
            try:
                WebUtils.js_click(self.driver, WebDriverWait(self.driver, 6).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Enviar') or contains(.,'Submit')]"))
                ))
            except Exception:
                email_pbi.send_keys(Keys.RETURN)
        except: pass
        try:
            WebDriverWait(self.driver, 10).until(EC.url_contains("login.microsoftonline"))
            mail = WebDriverWait(self.driver, 8).until(EC.element_to_be_clickable((By.NAME, "loginfmt")))
            WebUtils.safe_type_fast(mail, PBI_LOGIN_EMAIL)
            WebUtils.js_click(self.driver, WebDriverWait(self.driver, 6).until(EC.element_to_be_clickable((By.ID, "idSIButton9"))))
        except: pass

        self.driver.switch_to.default_content()
        try:
            pwd = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='i0118']")))
            WebUtils.safe_type_fast(pwd, PBI_SENHA)
            WebUtils.js_click(self.driver, WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, "idSIButton9"))))
        except: pass
        try:
            WebUtils.js_click(self.driver, WebDriverWait(self.driver, 6).until(EC.element_to_be_clickable((By.ID, "idBtn_Back"))))
        except: pass
        WebDriverWait(self.driver, 60).until(EC.url_contains("app.powerbi.com"))

    def enter_report_context(self, hard_timeout=70):
        end = time.time() + hard_timeout
        while time.time() < end:
            self.driver.switch_to.default_content()
            if self.driver.find_elements(By.CSS_SELECTOR, "#pvExplorationHost, div.visual, div.canvas, [role='presentation']"):
                time.sleep(5)
                return
            for f in self.driver.find_elements(By.TAG_NAME, "iframe"):
                try:
                    self.driver.switch_to.default_content()
                    self.driver.switch_to.frame(f)
                    if self.driver.find_elements(By.CSS_SELECTOR, "#pvExplorationHost, div.visual, div.canvas, [role='presentation']"):
                        time.sleep(5)
                        return
                except: pass
            time.sleep(1)
        raise TimeoutError("Report não ficou pronto (iframe/main).")

    def aplicar_filtro_coluna_exata(self, franquias):
        print("🔍 Buscando barra de pesquisa EXATA da 'Franquia'...")
        titulo_el = self.wait.until(EC.visibility_of_element_located((By.XPATH, "//*[normalize-space(text())='Franquia' or @title='Franquia']")))
        inputs_pesquisa = self.driver.find_elements(
            By.XPATH,
            "//input[contains(@placeholder,'Pesquisar') or contains(@aria-label,'Pesquisar')"
            " or contains(@placeholder,'Search') or contains(@aria-label,'Search')]",
        )
        inputs_visiveis = [inp for inp in inputs_pesquisa if inp.is_displayed()]

        search_input = min(inputs_visiveis, key=lambda inp: abs(inp.location['x'] - titulo_el.location['x'])) if inputs_visiveis else None
        if not search_input: raise RuntimeError("Pesquisa da Franquia não encontrada.")

        actions = ActionChains(self.driver)
        for franquia in franquias:
            self.driver.execute_script("arguments[0].value = '';", search_input)
            search_input.click()
            search_input.send_keys(franquia)
            item_xpath = f"//div[contains(@class,'slicerItemContainer')]//*[contains(text(),'{franquia}') or contains(@title,'{franquia}')]"
            try: WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.XPATH, item_xpath)))
            except TimeoutException: pass

            itens_visiveis = [i for i in self.driver.find_elements(By.XPATH, item_xpath) if i.is_displayed()]
            if itens_visiveis:
                target_item = min(itens_visiveis, key=lambda i: abs(i.location['x'] - search_input.location['x']))
                actions.key_down(Keys.CONTROL).click(target_item).key_up(Keys.CONTROL).perform()
                print(f"✅ Franquia flegada: {franquia}")
            else:
                print(f"⚠️ Franquia '{franquia}' não encontrada.")

        self.driver.execute_script("arguments[0].value = '';", search_input)
        search_input.click()
        search_input.send_keys(Keys.ESCAPE)

    def aplicar_filtro_franquia_popup(self, franquias):
        print("Abrindo popup de Franquia (relatório Relacionamento)...")
        if not self._open_franquia_combobox():
            raise RuntimeError("Não foi possível encontrar o combobox do filtro Franquia.")

        popup = WebDriverWait(self.driver, 12).until(EC.visibility_of_element_located((By.XPATH, self.POPUP_XPATH)))
        self._clear_popup(popup)
        self._select_popup_mult(popup, franquias)

        try:
            aplicar = popup.find_element(By.XPATH, ".//button[contains(.,'Aplicar') or contains(.,'Apply')]")
            WebUtils.js_click(self.driver, aplicar)
        except: pass
        print("Filtro de Franquia aplicado via popup.")

    def aplicar_filtro_data_filiacao(self, data_inicio):
        print(f"📅 Digitando data de filiação: {data_inicio}...")
        try:
            inp = self.wait.until(EC.presence_of_element_located((By.XPATH, self.XPATH_DATA_INPUT)))
        except TimeoutException:
            inp = self.wait.until(EC.presence_of_element_located((By.XPATH, "(//input[contains(@class,'date-slicer-input') and contains(@aria-label,'Data de início')])[1]")))

        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        time.sleep(0.2)
        inp.click()
        inp.send_keys(Keys.CONTROL, 'a')
        time.sleep(0.1)

        js = ("var el=arguments[0], v=arguments[1]; var s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;"
              "s.call(el,v); el.dispatchEvent(new Event('input',{bubbles:true})); el.dispatchEvent(new Event('change',{bubbles:true}));")
        self.driver.execute_script(js, inp, data_inicio)
        time.sleep(0.3)

        try:
            backdrop = self.driver.find_element(By.CSS_SELECTOR, "div.cdk-overlay-backdrop")
            if backdrop.is_displayed():
                ActionChains(self.driver).send_keys(Keys.ESCAPE).perform()
                time.sleep(0.3)
        except: pass

        self.driver.execute_script("arguments[0].click();", inp)
        time.sleep(0.1)
        inp.send_keys(Keys.RETURN)
        time.sleep(0.3)
        inp.send_keys(Keys.TAB)
        print(f"⏳ Aguardando 6s para atualização do BI...")
        time.sleep(6)
        print(f"✅ Data definida: {data_inicio}.")

    def download_report(self, base_name: str, titulo: str = "Filiados Ativos") -> Path:
        visual = self._find_table_visual(titulo)

        # ============================================================
        # RELATÓRIO 3 (BASE_VENDAS): Fluxo via cabeçalho Email + vcMenuBtn
        # ===========================================================
        if base_name == "BASE_VENDAS":
            if not self._open_more_menu_vendas(visual):
                raise RuntimeError(f"Não consegui abrir o menu 'Mais opções' da tabela para {base_name}.")

            # Clicar em "Exportar dados" no menu
            print("📤 Clicando em 'Exportar dados'...")
            try:
                export_item = WebDriverWait(self.driver, 8).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//div[contains(@class,'pbi-menu-item-text-container')]"
                        "//span[text()='Exportar dados' or text()='Export data']"
                    ))
                )
                WebUtils.js_click(self.driver, export_item)
            except TimeoutException:
                export_item = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((
                        By.CSS_SELECTOR,
                        "button[data-testid='pbimenu-item.Exportar dados'],"
                        " button[data-testid='pbimenu-item.Export data']",
                    ))
                )
                WebUtils.js_click(self.driver, export_item)

            click_ts = time.time()

            # Clicar no botão "Exportar" do diálogo
            print("💾 Clicando no botão 'Exportar' do diálogo...")
            try:
                export_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='export-btn']"))
                )
                WebUtils.js_click(self.driver, export_btn)
            except TimeoutException:
                export_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='Exportar' or normalize-space(.)='Export' or contains(@aria-label,'Exportar') or contains(@aria-label,'Export')]"))
                )
                WebUtils.js_click(self.driver, export_btn)

        # ============================================================
        # BLOCO ORIGINAL INTOCÁVEL (Relatórios 1 e 2)
        # ===========================================================
        else:
            if not self._open_more_menu(visual):
                raise RuntimeError(f"Não consegui abrir o menu 'Mais opções' da tabela para {base_name}.")

            try:
                export_item = WebDriverWait(self.driver, 8).until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "button[data-testid='pbimenu-item.Exportar dados'],"
                    " button[data-testid='pbimenu-item.Export data']",
                )))
            except TimeoutException:
                export_item = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((
                    By.XPATH,
                    "//*[normalize-space(text())='Exportar dados' or normalize-space(text())='Export data']",
                )))
            WebUtils.js_click(self.driver, export_item)
            click_ts = time.time()

            try:
                export_btn = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='export-btn']")))
            except TimeoutException:
                export_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='Exportar' or normalize-space(.)='Export' or contains(@aria-label,'Exportar') or contains(@aria-label,'Export')]")))

            WebUtils.js_click(self.driver, export_btn)

        # ============================================================
        # DOWNLOAD E RENOMEAÇÃO COMPARTILHADO
        # ===========================================================
        print(f"📥 Iniciando exportação: {base_name}")

        arquivo_baixado = FileProcessor.wait_for_download(DOWNLOAD_DIR, since_ts=click_ts)

        data_hoje = datetime.now().strftime("%d-%m-%Y")
        nome_novo = f"{base_name}_{data_hoje}{arquivo_baixado.suffix}"
        caminho_novo = arquivo_baixado.parent / nome_novo

        if caminho_novo.exists(): caminho_novo.unlink()
        arquivo_baixado.rename(caminho_novo)
        print(f"✅ Arquivo salvo e renomeado: {nome_novo}")

        return caminho_novo

    # --- Helpers Internos do Power BI ---
    def _open_franquia_combobox(self):
        wait = WebDriverWait(self.driver, 15)
        selectors = [(By.XPATH, "//div[@role='combobox' and contains(@aria-label,'Franquia')]"),
                     (By.CSS_SELECTOR, "[role='combobox'][aria-label*='Franquia']")]
        for by, sel in selectors:
            try:
                cb = wait.until(EC.element_to_be_clickable((by, sel)))
                WebUtils.js_click(self.driver, cb)
                return True
            except: pass

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(normalize-space(.),'Franquia')]")))
            cb = wait.until(EC.element_to_be_clickable((By.XPATH, "(.///*[contains(normalize-space(.),'Franquia')])[1]/following::*[@role='combobox' or self::button][1]")))
            WebUtils.js_click(self.driver, cb)
            return True
        except: pass
        return False

    def _clear_popup(self, popup):
        for label in ("Limpar selecoes", "Limpar selecao", "Limpar", "Desmarcar tudo", "Clear selections", "Clear selection", "Clear", "Deselect all"):
            try:
                btn = popup.find_element(By.XPATH, f".//button[contains(.,'{label}')]")
                WebUtils.js_click(self.driver, btn); time.sleep(0.3); return True
            except: continue
        return False

    def _select_popup_mult(self, popup, nomes):
        for nome in nomes:
            try: search = popup.find_element(By.CSS_SELECTOR, "input[aria-label*='Pesquisar'], input[aria-label*='Search'], input.searchInput")
            except: search = None

            if search: WebUtils.safe_type(search, nome, delay=0.01)

            try:
                item = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, f"{self.POPUP_XPATH}//div[contains(@class,'slicerItemContainer')][.//span[contains(@class,'slicerText') and contains(normalize-space(.),'{nome}')]]"
                )))
                target = item.find_element(By.CSS_SELECTOR, ".slicerCheckbox, .checkbox, .checkboxOutline")
            except:
                if search: WebUtils.safe_type(search, "")
                continue

            ActionChains(self.driver).key_down(Keys.CONTROL).move_to_element(target).pause(0.05).click(target).key_up(Keys.CONTROL).perform()
            if search: WebUtils.safe_type(search, ""); time.sleep(0.15)

    def _find_table_visual(self, titulo="Filiados Ativos"):
        try:
            title_xpath = f"//*[normalize-space(.)='Relatório Matrículas inadimplentes' or normalize-space(.)='{titulo}']"
            title = self.wait.until(EC.presence_of_element_located((By.XPATH, title_xpath)))
            return title.find_element(By.XPATH, "ancestor::*[contains(@data-automationid,'visual') or contains(@class,'visual')][1]")
        except: raise RuntimeError(f"Não achei a tabela '{titulo}'.")

    # ============================================================
    # MENU ORIGINAL INTOCÁVEL (Apenas para Relatórios 1 e 2)
    # ===========================================================
    def _open_more_menu(self, visual):
        actions = ActionChains(self.driver)
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", visual)
        time.sleep(0.1)
        try: actions.move_to_element(visual).pause(0.1).move_to_element_with_offset(visual, 30, 30).click().pause(0.1).perform()
        except: pass

        success = self.driver.execute_script("""
            const vis = arguments[0]; const r = vis.getBoundingClientRect();
            const xs = [10, 22, 36, 48]; const ys = [10, 14, 18, 22];
            for (let dx of xs){ for (let dy of ys){
                const el = document.elementFromPoint(Math.max(1, r.right - dx), r.top + dy);
                if (el){
                  el.dispatchEvent(new MouseEvent('mouseover', {bubbles:true})); el.dispatchEvent(new MouseEvent('mousemove', {bubbles:true})); el.click();
                  if (document.querySelector("pbi-menu[data-testid='pbi-menu']")) return true;
                }
            }} return false;
        """, visual)
        if success: return True

        try:
            actions.move_to_element(visual).click().key_down(Keys.SHIFT).key_down(Keys.F10).key_up(Keys.F10).key_up(Keys.SHIFT).perform()
            WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "pbi-menu[data-testid='pbi-menu']")))
            return True
        except: return False

    # ============================================================
    # MENU EXCLUSIVO PARA RELATÓRIO 3 (BASE_VENDAS)
    # CORREÇÃO StaleElementReferenceException:
    #   - _fresh_email_header() re-busca o elemento a cada uso
    #   - Loop de retentativas com captura de StaleElementReferenceException
    #   - visual também é validado/re-buscado antes de passar ao JS
    # ===========================================================
    def _open_more_menu_vendas(self, visual):
        actions = ActionChains(self.driver)

        print("📧 Localizando cabeçalho da coluna 'Email'...")

        hover_ok = False
        for attempt in range(3):
            eh = self._fresh_email_header()
            if eh is None:
                print("⚠️ Cabeçalho Email não encontrado.")
                break
            try:
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center', inline:'end'});", eh
                )
                time.sleep(0.6)

                eh = self._fresh_email_header()
                if eh is None:
                    break

                actions.move_to_element(eh).pause(1.0).perform()
                print("✅ Hover no cabeçalho Email realizado.")
                time.sleep(0.8)
                hover_ok = True
                break

            except StaleElementReferenceException:
                print(f"⚠️ Stale no email_header (tentativa {attempt + 1}/3) — re-buscando...")
                time.sleep(0.5)
                continue
            except Exception as e:
                print(f"⚠️ Erro inesperado no hover do email_header: {e}")
                break

        if not hover_ok:
            print("⚠️ Continuando sem hover no cabeçalho Email.")

        print("🔍 Buscando vcMenuBtn dentro do visual container correto...")

        eh_fresh = self._fresh_email_header()

        vis_fresh = None
        try:
            _ = visual.tag_name
            vis_fresh = visual
        except StaleElementReferenceException:
            print("⚠️ Visual stale — re-buscando pelo título 'Filiados Inativos'...")
            try:
                title = self.driver.find_element(By.XPATH, "//*[normalize-space(.)='Filiados Inativos']")
                vis_fresh = title.find_element(
                    By.XPATH, "ancestor::*[contains(@data-automationid,'visual') or contains(@class,'visual')][1]"
                )
            except Exception:
                vis_fresh = None

        btn = self.driver.execute_script("""
            var emailHdr = arguments[0];
            var vis      = arguments[1];

            function visibleBtn(root) {
                if (!root) return null;
                var b = root.querySelector(
                    "button.vcMenuBtn[data-testid='visual-more-options-btn']"
                );
                if (b) {
                    var r = b.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0) return b;
                }
                return null;
            }

            if (emailHdr) {
                var el = emailHdr;
                for (var i = 0; i < 30 && el && el !== document.body; i++) {
                    el = el.parentElement;
                    var found = visibleBtn(el);
                    if (found) return found;
                }
            }

            if (vis) {
                var el2 = vis;
                for (var j = 0; j < 10 && el2 && el2 !== document.body; j++) {
                    var found2 = visibleBtn(el2);
                    if (found2) return found2;
                    el2 = el2.parentElement;
                }
            }

            if (emailHdr) {
                var er = emailHdr.getBoundingClientRect();
                var all = Array.from(document.querySelectorAll(
                    "button.vcMenuBtn[data-testid='visual-more-options-btn']"
                )).filter(function(b) {
                    var r = b.getBoundingClientRect();
                    return r.width > 0 && r.height > 0;
                });
                if (all.length) {
                    return all.reduce(function(best, b) {
                        var r = b.getBoundingClientRect();
                        var d = Math.pow(r.left - er.right, 2) + Math.pow(r.top - er.top, 2);
                        return d < best.d ? {b: b, d: d} : best;
                    }, {b: all[0], d: Infinity}).b;
                }
            }

            return null;
        """, eh_fresh, vis_fresh)

        if btn:
            try:
                WebUtils.js_click(self.driver, btn)
                print("✅ vcMenuBtn clicado (DOM traversal).")
                time.sleep(0.5)
                WebDriverWait(self.driver, 6).until(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Exportar dados') or contains(text(),'Export data')]"))
                )
                return True
            except Exception as e:
                print(f"⚠️ JS encontrou o btn mas clique falhou: {e}")

        print("🔄 Fallback Python: aguardando vcMenuBtn clicável...")
        for attempt, (by, sel) in enumerate([
            (By.XPATH,        "//button[contains(@class,'vcMenuBtn') and @data-testid='visual-more-options-btn']"),
            (By.XPATH,        "//button[@data-testid='visual-more-options-btn' and @aria-haspopup='true']"),
            (By.CSS_SELECTOR, "button.vcMenuBtn[data-testid='visual-more-options-btn']"),
        ], 1):
            try:
                b = WebDriverWait(self.driver, 6).until(EC.element_to_be_clickable((by, sel)))
                WebUtils.js_click(self.driver, b)
                print(f"✅ vcMenuBtn clicado via fallback Python #{attempt}.")
                time.sleep(0.5)
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Exportar dados') or contains(text(),'Export data')]"))
                )
                return True
            except Exception:
                continue

        print("❌ Todas as tentativas falharam para _open_more_menu_vendas.")
        return False


# ==========================================
# AUTOMAÇÃO: REDE SERVICE
# ==========================================
class RedeServiceBot:
    # O formulário de importação (o que o botão "Novo" abre) tem URL própria.
    # Indicada pelo usuário: depois de logado, essa URL já cai direto no
    # formulário, sem depender do clique no "Novo" — que é justamente o passo
    # que mais deu problema (botão "clicável" antes do binding do Angular
    # terminar, modal que não abria, e a recuperação disso refazendo login no
    # meio do fluxo).
    URL_IMPORTACAO_NOVA = "https://cobranca01.redeservice.com.br/cobranca.be.cartaotodos/Importacao/Incluir"

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)

    def login_and_navigate(self):
        print("🌐 Iniciando importação no RedeService...")
        self.driver.get(RS_URL_IMPORT)
        # _fazer_login() já espera o pós-login terminar de verdade (sair da
        # tela de login) antes de seguir. Sem isso, se o redirecionamento
        # demorar um pouco mais, a navegação seguinte roda ainda em cima da
        # tela de login e nada do que vem depois encontra o que precisa.
        self._fazer_login()
        print("✅ Logado no RedeService.")

    def _esta_na_tela_login(self) -> bool:
        return bool(self.driver.find_elements(By.ID, "Login")) and bool(self.driver.find_elements(By.ID, "PasswordTextBox"))

    def _abrir_formulario_importacao(self):
        """
        Abre o formulário de importação indo direto na URL .../Importacao/Incluir.

        Antes isso era feito clicando no botão "Novo" (`demo-btn-addrow`) e
        esperando o modal abrir — e era aí que a automação mais emperrava (o
        botão aparecia "clicável" antes do Angular terminar o binding, o
        modal não abria, e a recuperação disso acabava refazendo login no
        meio do fluxo). Ir direto pela URL, com a sessão já estabelecida,
        pula esse passo inteiro.
        """
        for tentativa in range(2):
            self.driver.get(self.URL_IMPORTACAO_NOVA)

            if self._esta_na_tela_login():
                print("⚠️ A sessão caiu de volta para a tela de login do RedeService; refazendo login do zero...")
                self.login_limpo()
                self.driver.get(self.URL_IMPORTACAO_NOVA)

            try:
                WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID, "ddlTipoImportacao")))
                return
            except TimeoutException:
                if tentativa == 0:
                    print("⚠️ Formulário de importação não abriu pela URL direta; refazendo login do zero e tentando de novo...")
                    self.login_limpo()
        raise TimeoutException(
            f"Formulário de importação (ddlTipoImportacao) não abriu em {self.URL_IMPORTACAO_NOVA} após 2 tentativas."
        )

    def login_limpo(self):
        """
        Descarta a sessão atual (cookies + storage) e loga de novo, do zero.

        Observação de produção, consistente em várias execuções: o
        RedeService invalida a sessão do robô logo depois de uma importação
        ser enviada. A PRIMEIRA base — que roda logo após um login novo —
        sempre passa; as seguintes, que reaproveitavam a sessão já usada,
        caíam na tela de login. Em vez de tentar preservar uma sessão que o
        servidor já considera morta, cada base começa com um login limpo,
        recriando a condição que comprovadamente funciona.
        """
        try:
            self.driver.delete_all_cookies()
            self.driver.execute_script("try { localStorage.clear(); sessionStorage.clear(); } catch (e) {}")
        except Exception:
            pass
        self.driver.get(RS_URL_IMPORT)
        self._fazer_login()

    def _fazer_login(self):
        """Preenche e envia o formulário de login, esperando o pós-login terminar."""
        WebUtils.safe_type(self.wait.until(EC.element_to_be_clickable((By.ID, "Login"))), RS_LOGIN)
        WebUtils.safe_type(self.wait.until(EC.element_to_be_clickable((By.ID, "PasswordTextBox"))), RS_SENHA)
        WebUtils.js_click(self.driver, self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']"))))
        WebDriverWait(self.driver, 30).until(lambda d: not self._esta_na_tela_login())

    def _selecionar(self, select_id: str, value: str):
        """
        Seleciona um <option> e dispara change/input explicitamente.

        O select de Layout é populado quando os campos "pai" (Tipo de
        Importação/Cliente) mudam. Nem sempre a seleção via Selenium faz o
        framework da página perceber a mudança — sem o evento, o Layout
        nunca chega a ser carregado e a espera por ele estoura.
        """
        el = self.wait.until(EC.element_to_be_clickable((By.ID, select_id)))
        SeleniumSelect(el).select_by_value(value)
        self.driver.execute_script(
            "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));"
            "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
            el,
        )
        time.sleep(1)

    def _select_quando_disponivel(self, select_id: str, value: str, timeout: int = 40):
        """
        Seleciona um <option> por value, esperando ele existir antes.

        Alguns selects (como o de Layout) são populados dinamicamente depois
        que um select "pai" (Tipo de Importação/Cliente) muda — selecionar
        direto pode disparar NoSuchElementException porque a opção ainda não
        chegou no DOM.

        Se a opção não aparecer, despeja no log o que a página REALMENTE tem
        (todos os selects e suas opções) antes de estourar: já perdi tempo
        demais supondo o que estava na tela em vez de olhar.
        """
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda d: d.find_elements(By.CSS_SELECTOR, f"#{select_id} option[value='{value}']")
            )
        except TimeoutException:
            self._dump_selects_da_pagina(select_id, value)
            raise
        SeleniumSelect(self.driver.find_element(By.ID, select_id)).select_by_value(value)

    def _dump_selects_da_pagina(self, select_id: str, value: str):
        """Lista no log todos os <select> da página e suas opções (diagnóstico)."""
        print(f"⚠️ A opção value='{value}' não apareceu em #{select_id}. Campos que a página tem agora:")
        try:
            print(f"   URL atual: {self.driver.current_url}")
            for el in self.driver.find_elements(By.TAG_NAME, "select"):
                sid = el.get_attribute("id") or "(sem id)"
                nome = el.get_attribute("name") or "(sem name)"
                visivel = el.is_displayed()
                opcoes = [
                    f"{o.get_attribute('value')}={(o.text or '').strip()}"
                    for o in el.find_elements(By.TAG_NAME, "option")
                ]
                print(f"   <select id={sid} name={nome} visível={visivel}> opções: {opcoes[:25]}")
        except Exception as e:
            print(f"   (não consegui listar os selects: {e})")

    def _upload_dropzone(self, caminho_csv: Path):
        """
        Entrega o arquivo ao Dropzone de verdade.

        CAUSA RAIZ do "importa mas não importa": escrever o caminho no
        <input type="file"> escondido faz o arquivo chegar no input, mas o
        Dropzone (biblioteca JS) NÃO fica sabendo — ele não lê o input, ele
        reage a eventos. Resultado: o script achava que tinha subido o
        arquivo, o RedeService recebia o formulário vazio e respondia
        "Informe o(s) arquivo(s) para importação!" (confirmado no log, no
        texto da página depois do Enviar), sem nada ser importado.

        A correção é disparar o evento 'change' no input depois do
        send_keys, e — para o caso do Dropzone só escutar drag-and-drop —
        também despachar um evento 'drop' sintético com o arquivo no
        dataTransfer.
        """
        self.driver.execute_script("""
            document.querySelectorAll('.dropzone input[type="file"], input.dz-hidden-input').forEach(function(el) {
                el.style.display = 'block'; el.style.opacity = '1'; el.style.position = 'relative'; el.style.width = '1px'; el.style.height = '1px';
            });
        """)
        file_input = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.dropzone input[type="file"], input.dz-hidden-input')))
        file_input.send_keys(str(caminho_csv.resolve()))

        # Sem isso o Dropzone ignora o arquivo (ver docstring).
        # O input é re-consultado DENTRO do JS: passar a referência Python de
        # volta pro script estourava StaleElementReferenceException, porque o
        # Dropzone recria esse elemento assim que recebe o arquivo.
        self.driver.execute_script("""
            var input = document.querySelector('.dropzone input[type="file"], input.dz-hidden-input');
            if (!input) { return; }
            input.dispatchEvent(new Event('change', {bubbles: true}));

            // Se o Dropzone só escuta drop, replica o arrastar-e-soltar
            // usando os arquivos que já estão no input.
            var zona = document.querySelector('.dropzone');
            if (zona && input.files && input.files.length) {
                var dt = new DataTransfer();
                for (var i = 0; i < input.files.length; i++) { dt.items.add(input.files[i]); }
                ['dragenter', 'dragover', 'drop'].forEach(function (nome) {
                    var ev = new DragEvent(nome, {bubbles: true, cancelable: true, dataTransfer: dt});
                    zona.dispatchEvent(ev);
                });
            }
        """)
        print(f"📎 Arquivo entregue ao Dropzone: {caminho_csv.name}")

        # Confirma que o Dropzone realmente registrou o arquivo (ele mostra o
        # nome na tela quando aceita). Sem essa checagem voltamos a "achar"
        # que subiu.
        try:
            WebDriverWait(self.driver, 30).until(
                lambda d: caminho_csv.name in d.find_element(By.TAG_NAME, "body").text
            )
            print("✅ Dropzone confirmou o arquivo na tela.")
        except TimeoutException:
            print("⚠️ O nome do arquivo NÃO apareceu na tela — o Dropzone pode não ter aceitado.")

        self._garantir_upload_concluido(caminho_csv)

    def _garantir_upload_concluido(self, caminho_csv: Path, timeout: int = 180):
        """
        Espera o Dropzone terminar de ENVIAR o arquivo pro servidor.

        O arquivo aparecer na tela só significa que entrou na fila do
        Dropzone — não que subiu. Em produção o RedeService continuava
        respondendo "Informe o(s) arquivo(s) para importação!" mesmo com o
        nome do arquivo visível: a fila estava montada, o upload não tinha
        acontecido.

        Aqui consultamos a API do próprio Dropzone (`Dropzone.forElement`),
        que expõe o status real de cada arquivo ('added'/'queued',
        'uploading', 'success', 'error'). Se estiver parado na fila,
        mandamos processar (`processQueue()`), que é justamente o que a
        página faria se `autoProcessQueue` estivesse ligado.
        """
        estado_js = """
            if (!window.Dropzone || !Dropzone.forElement) { return 'sem-dropzone'; }
            var zona = document.querySelector('.dropzone');
            if (!zona) { return 'sem-zona'; }
            var dz;
            try { dz = Dropzone.forElement(zona); } catch (e) { return 'sem-instancia'; }
            if (!dz || !dz.files) { return 'sem-instancia'; }
            return dz.files.map(function (f) { return f.name + '=' + f.status; }).join(' | ') || 'sem-arquivos';
        """
        processar_js = """
            var zona = document.querySelector('.dropzone');
            var dz = Dropzone.forElement(zona);
            dz.processQueue();
        """

        estado = self.driver.execute_script(estado_js)
        print(f"🔎 Estado do Dropzone: {estado}")

        if estado in ("sem-dropzone", "sem-zona", "sem-instancia", "sem-arquivos"):
            # Sem a API disponível não dá pra acompanhar; o upload pode ser
            # feito no submit do formulário. Segue o fluxo — a checagem de
            # recusa depois do Enviar continua valendo como rede de segurança.
            print("ℹ️ Não consigo consultar a fila do Dropzone; seguindo assim mesmo.")
            return

        if "=queued" in estado or "=added" in estado:
            print("⏫ Arquivo parado na fila do Dropzone; mandando processar...")
            try:
                self.driver.execute_script(processar_js)
            except Exception as e:
                print(f"⚠️ processQueue() falhou: {e}")

        fim = time.time() + timeout
        while time.time() < fim:
            estado = self.driver.execute_script(estado_js)
            if "=success" in estado:
                print(f"✅ Upload concluído no Dropzone: {estado}")
                return
            if "=error" in estado:
                raise RuntimeError(f"O Dropzone marcou o upload como erro: {estado}")
            time.sleep(3)

        raise RuntimeError(
            f"O upload não terminou em {timeout}s (estado do Dropzone: {estado}). "
            f"Enviar agora só geraria 'Informe o(s) arquivo(s) para importação'."
        )

    # Unidade "AGIL SOLUÇÕES INTEGRADAS", conforme o passo a passo do usuário.
    # No modal antigo ela já vinha marcada por padrão (selected="selected"), por
    # isso o script nunca precisou mexer nela. Na página /Importacao/Incluir isso
    # pode não valer — e um Enviar com Unidade vazia é candidato direto a
    # "o script diz que enviou mas nada é importado".
    UNIDADE_VALUE = "000001"

    def _selecionar_unidade(self):
        """
        Preenche o campo Unidade, achando o <select> pelo value da opção.

        Não tenho o id/name desse campo (o usuário mandou só o <option>), então
        procuro entre os selects da página um que tenha a opção 000001 e que
        não seja um dos que já conheço.
        """
        conhecidos = {"ddlTipoImportacao", "ddlcliente", "ddllayout"}
        for el in self.driver.find_elements(By.TAG_NAME, "select"):
            if (el.get_attribute("id") or "") in conhecidos:
                continue
            opcoes = el.find_elements(By.CSS_SELECTOR, f"option[value='{self.UNIDADE_VALUE}']")
            if not opcoes:
                continue
            sel = SeleniumSelect(el)
            atual = sel.first_selected_option.get_attribute("value") if sel.all_selected_options else None
            if atual == self.UNIDADE_VALUE:
                print(f"ℹ️ Unidade já vem selecionada ({self.UNIDADE_VALUE}); não mexo.")
                return
            sel.select_by_value(self.UNIDADE_VALUE)
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));"
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                el,
            )
            print(f"✅ Unidade selecionada ({self.UNIDADE_VALUE}) em #{el.get_attribute('id') or '(sem id)'}.")
            time.sleep(1)
            return
        print("⚠️ Não achei nenhum campo de Unidade com a opção 000001 nessa tela.")

    def _dump_estado_do_formulario(self, momento: str):
        """Loga o que cada campo do formulário tem selecionado de verdade."""
        print(f"🔎 Estado do formulário ({momento}):")
        try:
            for el in self.driver.find_elements(By.TAG_NAME, "select"):
                if not el.is_displayed():
                    continue
                sel = SeleniumSelect(el)
                escolhido = sel.first_selected_option.get_attribute("value") if sel.all_selected_options else "(nada)"
                texto = sel.first_selected_option.text.strip() if sel.all_selected_options else ""
                sid = el.get_attribute("id") or el.get_attribute("name") or "(sem id)"
                print(f"   #{sid} = '{escolhido}' ({texto})")
        except Exception as e:
            print(f"   (não consegui ler os campos: {e})")

    def _dump_texto_da_pagina(self, momento: str, limite: int = 1200):
        """
        Loga o texto visível da página.

        O RedeService não dá confirmação de sucesso, mas se o formulário for
        rejeitado por validação (campo obrigatório vazio, por exemplo) a
        mensagem aparece na tela — e até agora o script nunca olhou pra ela.
        """
        print(f"🔎 Texto da página ({momento}):")
        try:
            print(f"   URL: {self.driver.current_url}")
            texto = self.driver.find_element(By.TAG_NAME, "body").text
            print("   " + (texto[:limite] or "(vazio)").replace("\n", "\n   "))
        except Exception as e:
            print(f"   (não consegui ler o texto: {e})")

    def importar_base(self, layout_value: str, caminho_csv: Path, is_first: bool = False):
        # Cada base começa com uma sessão nova. O RedeService invalida a
        # sessão logo depois de uma importação ser enviada — a primeira base
        # sempre passava (login novo), as seguintes caíam na tela de login.
        # Logar do zero antes de cada uma recria a condição que funciona.
        if not is_first:
            print("🔄 Refazendo login antes da próxima base (a sessão não sobrevive a uma importação)...")
            self.login_limpo()

        # Vai direto pro formulário pela URL — não precisa passar pela grade
        # e clicar em "Novo".
        self._abrir_formulario_importacao()
        self._selecionar("ddlTipoImportacao", "11")
        self._selecionar_unidade()
        self._selecionar("ddlcliente", "000003")
        self._select_quando_disponivel("ddllayout", layout_value)

        self._upload_dropzone(caminho_csv)
        print("⏳ Aguardando processamento do upload (10s)...")
        time.sleep(10)

        self._dump_estado_do_formulario("ANTES do Enviar")
        WebUtils.js_click(self.driver, self.wait.until(EC.element_to_be_clickable((By.ID, "btnEnviar"))))
        print(f"📤 Enviar (Layout {layout_value}) clicado.")
        time.sleep(5)
        self._dump_texto_da_pagina("DEPOIS do Enviar")
        self._conferir_rejeicao(layout_value)

    # Mensagens de validação que o RedeService mostra quando recusa o envio.
    # "Informe o(s) arquivo(s) para importação!" foi a que apareceu em
    # produção enquanto o Dropzone não recebia o arquivo de verdade — e o
    # script seguia dizendo "enviado com sucesso" por cima dela.
    MENSAGENS_DE_REJEICAO = (
        "Informe o(s) arquivo(s) para importação",
        "campo obrigatório",
        "Campo obrigatório",
    )

    def _conferir_rejeicao(self, layout_value: str):
        """Falha se a tela mostrar mensagem de validação depois do Enviar."""
        try:
            texto = self.driver.find_element(By.TAG_NAME, "body").text
        except Exception:
            texto = ""
        for msg in self.MENSAGENS_DE_REJEICAO:
            if msg in texto:
                raise RuntimeError(
                    f"O RedeService recusou a importação (Layout {layout_value}): '{msg}'. "
                    f"Nada foi importado."
                )
        print(f"🚀 Importação (Layout {layout_value}) enviada sem mensagem de recusa.")


# ==========================================
# FLUXO DE EXECUÇÃO PRINCIPAL
# ==========================================
def main():
    if not PBI_LOGIN_EMAIL or not PBI_SENHA:
        raise RuntimeError("Defina as variáveis de ambiente PBI_LOGIN_EMAIL e PBI_SENHA (GitHub Secrets).")
    if not RS_LOGIN or not RS_SENHA:
        raise RuntimeError("Defina as variáveis de ambiente RS_LOGIN e RS_SENHA (GitHub Secrets).")

    arquivo_csv_inadimplencia = None
    arquivo_relacionamento = None
    arquivo_vendas = None
    contagem_inadimplencia = 0
    contagem_relacionamento = 0
    contagem_vendas = 0

    # --- ETAPA 1: DOWNLOADS POWER BI ---
    driver_pbi = BrowserFactory.create_chrome()
    try:
        pbi_bot = PowerBIBot(driver_pbi)

        print("\n========== RELATÓRIO 1: INADIMPLÊNCIA ==========")
        driver_pbi.get(URL_INADIMPLENCIA)
        pbi_bot.login()
        pbi_bot.enter_report_context()
        pbi_bot.aplicar_filtro_coluna_exata(FRANQUIAS)
        base_1_xlsx = pbi_bot.download_report("BASE_INADIMPLENCIA")
        arquivo_csv_inadimplencia, contagem_inadimplencia = FileProcessor.processar_planilha_inadimplencia(base_1_xlsx)

        print("\n========== RELATÓRIO 2: RELACIONAMENTO ==========")
        driver_pbi.get(URL_RELACIONAMENTO)
        pbi_bot.enter_report_context()
        pbi_bot.aplicar_filtro_franquia_popup(FRANQUIAS)
        pbi_bot.aplicar_filtro_data_filiacao(DATA_FILIACAO_INICIO)
        arquivo_relacionamento = pbi_bot.download_report("BASE_RELACIONAMENTO")
        contagem_relacionamento = FileProcessor.contar_linhas(arquivo_relacionamento)

        print("\n========== RELATÓRIO 3: VENDAS ==========")
        driver_pbi.get(URL_VENDAS)
        pbi_bot.enter_report_context()
        pbi_bot.aplicar_filtro_coluna_exata(FRANQUIAS_VENDAS)
        arquivo_vendas = pbi_bot.download_report("BASE_VENDAS", titulo="Filiados Inativos")
        contagem_vendas = FileProcessor.contar_linhas(arquivo_vendas)

    except Exception as e:
        nome_print = f"erro_pbi_{datetime.now().strftime('%H%M%S')}.png"
        print(f"\n❌ [ERRO] Falha na etapa Power BI: {e}\n")
        try:
            driver_pbi.save_screenshot(nome_print)
            print(f"Screenshot salva em: {os.path.abspath(nome_print)}")
            with open(nome_print.replace(".png", ".html"), "w", encoding="utf-8") as f:
                f.write(driver_pbi.page_source)
            print(f"HTML da página salvo em: {os.path.abspath(nome_print.replace('.png', '.html'))}")
        except Exception as e2:
            print(f"Não consegui salvar screenshot/HTML de diagnóstico: {e2}")
        raise
    finally:
        print("Fechando navegador do PBI...")
        driver_pbi.quit()

    # Backup no Google Drive dos arquivos extraídos (best-effort). Direto na
    # pasta de destino, sem subpasta por data — se já existir um arquivo com
    # o mesmo nome, ele é substituído em vez de duplicado.
    for arquivo in (arquivo_csv_inadimplencia, arquivo_relacionamento, arquivo_vendas):
        if arquivo:
            upload_file(str(arquivo), GDRIVE_FOLDER_ID)

    # --- ETAPA 2: IMPORTAÇÃO REDESERVICE ---
    if not (arquivo_csv_inadimplencia and arquivo_relacionamento and arquivo_vendas):
        print("⚠️ Um ou mais arquivos falharam no download/processamento. Abortando upload.")
        raise RuntimeError("Falha no download/processamento de um ou mais relatórios do Power BI.")

    driver_rs = BrowserFactory.create_chrome()
    try:
        rs_bot = RedeServiceBot(driver_rs)
        rs_bot.login_and_navigate()

        print("\n--- Importação 1: BASE_INADIMPLENCIA ---")
        rs_bot.importar_base("79", arquivo_csv_inadimplencia, is_first=True)

        print("\n--- Importação 2: BASE_RELACIONAMENTO ---")
        rs_bot.importar_base("81", arquivo_relacionamento)

        print("\n--- Importação 3: BASE_VENDAS ---")
        rs_bot.importar_base("82", arquivo_vendas)

        with open("contagens.json", "w", encoding="utf-8") as f:
            json.dump({
                "inadimplencia": contagem_inadimplencia,
                "relacionamento": contagem_relacionamento,
                "vendas": contagem_vendas,
            }, f)
        print(f"\n📊 Linhas importadas — Inadimplência: {contagem_inadimplencia}, Relacionamento: {contagem_relacionamento}, Vendas: {contagem_vendas}")

    except Exception as e:
        nome_print = f"erro_rs_{datetime.now().strftime('%H%M%S')}.png"
        print(f"\n❌ [ERRO] Falha na etapa de importação RedeService: {e}\n")
        try:
            driver_rs.save_screenshot(nome_print)
            print(f"Screenshot salva em: {os.path.abspath(nome_print)}")
            with open(nome_print.replace(".png", ".html"), "w", encoding="utf-8") as f:
                f.write(driver_rs.page_source)
            print(f"HTML da página salvo em: {os.path.abspath(nome_print.replace('.png', '.html'))}")
        except Exception as e2:
            print(f"Não consegui salvar screenshot/HTML de diagnóstico: {e2}")
        raise
    finally:
        print("\n🏁 Todas as importações concluídas. Fechando navegador.")
        driver_rs.quit()

if __name__ == "__main__":
    main()
